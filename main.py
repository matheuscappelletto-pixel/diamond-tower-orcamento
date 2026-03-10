"""
DIAMOND TOWER — Automação Orçamentária via XLSX

Fluxo:
1. Procura um arquivo .xlsx em /entradas
2. Lê o extrato exportado da Guarida
3. Pega apenas os débitos
4. Classifica os lançamentos
5. Atualiza a planilha PF.xlsx
6. Salva notas/comentários nas células
7. Move o extrato para /processados
"""

import os
import re
import json
import shutil
import unicodedata
from datetime import datetime
from pathlib import Path

import anthropic
from openpyxl import load_workbook

# ── Configuração ──────────────────────────────────────────────

BASE_DIR = Path(__file__).resolve().parent
ENTRADAS_DIR = BASE_DIR / "entradas"
PROCESSADOS_DIR = BASE_DIR / "processados"
PF_PATH = BASE_DIR / "PF.xlsx"

ANTHROPIC_KEY = os.environ["ANTHROPIC_KEY"]

# ── Mapeamento mês → coluna na planilha ───────────────────────

MONTH_TO_COL = {
    (3, 2025): "F",
    (4, 2025): "G",
    (5, 2025): "H",
    (6, 2025): "I",
    (7, 2025): "J",
    (8, 2025): "K",
    (9, 2025): "L",
    (10, 2025): "M",
    (11, 2025): "N",
    (12, 2025): "O",
    (1, 2026): "P",
    (2, 2026): "Q",
    (3, 2026): "R",
}

# ── Mapeamento linhas da planilha ─────────────────────────────

ROW_MAP = {
    4: {"empresa": "ALTO PADRAO", "desc": "Alto Padrão — Mão de Obra",
        "keywords": ["alto padrao", "onx", "portaria", "recepcao", "monitoramento", "ronda", "porteiro", "limpeza"]},
    11: {"empresa": "CAPPELLETTO", "desc": "Cappelletto — Gestores presenciais",
         "keywords": ["cappelletto", "gestao", "servico auxiliar de administracao condominial"]},
    12: {"empresa": "GUARIDA", "desc": "Guarida — Auxiliar Administração",
         "keywords": ["guarida", "taxa de administracao", "taxa adm"]},
    13: {"empresa": "ORGANICO", "desc": "Orgânico — Oficial Manutenção 44h",
         "keywords": ["marco aurelio", "oficial manutencao", "salario marco aurelio"]},
    14: {"empresa": "ORGANICO", "desc": "Orgânico — Remuneração Subsíndico",
         "keywords": ["subsidico", "sub sindico", "pro-labore subsidico", "remuneracao subsidico"]},
    17: {"empresa": "FG PISCINAS", "desc": "FG Piscinas — Limpeza/Química",
         "keywords": ["fg piscinas", "piscina", "espelho d'agua", "espelho dagua"]},
    18: {"empresa": "ELITE", "desc": "Elite — Elevadores Atlas",
         "keywords": ["manut.elevador", "manut elevador", "atlas", "schindler", "elevadores atlas schindler"]},
    19: {"empresa": "STEMAC", "desc": "Stemac — Gerador",
         "keywords": ["stemac", "gerador", "manut. gerador"]},
    20: {"empresa": "BELLINI", "desc": "Bellini — Consultoria Jurídica",
         "keywords": ["bellini", "honorarios advocaticios", "juridica", "advocaticio"]},
    21: {"empresa": "AUDITORIA", "desc": "Auditoria Externa",
         "keywords": ["auditoria", "taborda", "pastas contabeis"]},
    24: {"empresa": "MULTIPLAN", "desc": "Multiplan — Lixo/Entulho",
         "keywords": ["multiplan", "lixo", "entulho"]},
    25: {"empresa": "ALPINISMO", "desc": "Alpinismo — Fachada",
         "keywords": ["fachada", "uperclean", "rest.fachada"]},
    26: {"empresa": "DEMANDA", "desc": "Manutenção Bombas Recalque",
         "keywords": ["bomba", "recalque", "manut.bomba dagua"]},
    27: {"empresa": "", "desc": "Manutenção Interfones",
         "keywords": ["interfone", "porteiro eletronico", "mactel"]},
    28: {"empresa": "", "desc": "Manutenção Predial",
         "keywords": ["hidrojateamento", "esgoto", "portao", "hidraulico", "desobstrucao", "manut.portao", "manutencao predial"]},
    29: {"empresa": "", "desc": "Incêndio",
         "keywords": ["incendio", "hidrante", "extintor"]},
    30: {"empresa": "", "desc": "Obrigações Legais",
         "keywords": ["laudo", "spda", "ppra", "pcmso", "qualidade do ar", "obrigacao legal", "ativa medicina", "toxilab"]},
    31: {"empresa": "", "desc": "Controle Pragas / Caixa D'água",
         "keywords": ["caixa dagua", "caixa d agua", "limpeza do reservatorio", "eco-ambiental", "dedetiz"]},
    32: {"empresa": "", "desc": "Paisagismo",
         "keywords": ["paisagismo"]},
    35: {"empresa": "", "desc": "Chaveiro",
         "keywords": ["chaveiro"]},
    36: {"empresa": "", "desc": "Lâmpadas",
         "keywords": ["lampada", "lampadas"]},
    37: {"empresa": "", "desc": "Obras Diversas",
         "keywords": ["obra", "mureta", "totens", "benfeitorias"]},
    38: {"empresa": "", "desc": "Material Ferragem/Elétrica/Hidráulica",
         "keywords": ["procal", "material", "mat. eletrico", "ferragens", "tintas", "material constr", "material pintura"]},
    39: {"empresa": "", "desc": "Material Expediente",
         "keywords": ["expediente", "escritorio", "impressora", "fortpel"]},
    40: {"empresa": "", "desc": "Material Limpeza",
         "keywords": ["material limpeza", "mat.limpeza", "descartaveis"]},
    41: {"empresa": "ELITE", "desc": "Elite — Peças Elevador",
         "keywords": ["elevador pecas", "peca elevador", "pecas elevador", "conserto elevador"]},
    42: {"empresa": "", "desc": "Móveis e Utensílios",
         "keywords": ["movel", "utensilio", "copa funcionarios"]},
    45: {"empresa": "", "desc": "Peças Gerador / Óleo Diesel",
         "keywords": ["bateria gerador", "diesel", "oleo diesel"]},
    49: {"empresa": "", "desc": "Telefonia/Internet",
         "keywords": ["vivo", "telefonica", "telefonia", "internet", "celular"]},
    51: {"empresa": "", "desc": "Impostos DIRF/DARF/ISS/PIS/COFINS",
         "keywords": ["issqn", "inss", "fgts", "secovimed", "efd-reinf", "retencao issqn", "declaracao mensal"]},
    52: {"empresa": "", "desc": "Assembleia",
         "keywords": ["assembleia", "reuniao de conselho", "almocos"]},
    59: {"empresa": "", "desc": "Água e Esgoto",
         "keywords": ["dmae", "consumo dagua", "agua", "esgoto"]},
    60: {"empresa": "", "desc": "Energia Elétrica",
         "keywords": ["ceee", "consumo luz", "energia eletrica"]},
    61: {"empresa": "", "desc": "Seguro",
         "keywords": ["seguro", "tokio marine"]},
    65: {"empresa": "", "desc": "Despesas Reembolsáveis",
         "keywords": ["reembolso", "ressarcimento", "uber", "reemb.mat"]},
    66: {"empresa": "", "desc": "Honorários Advocatícios",
         "keywords": ["honorario advocaticio", "honorarios advocaticios"]},
    68: {"empresa": "", "desc": "Transferências Entre Contas",
         "keywords": ["envio /transf", "transferencia entre contas", "aplicacao"]},
}

# ── Helpers ───────────────────────────────────────────────────

def norm(texto: str) -> str:
    t = unicodedata.normalize("NFKD", str(texto).lower())
    return "".join(c for c in t if not unicodedata.combining(c))


def formatar_brl(valor: float) -> str:
    return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def limpar_valor_excel(v) -> float:
    if v is None:
        return 0.0

    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip()
    if not s:
        return 0.0

    s = s.replace("R$", "").replace(" ", "")
    s = s.replace(".", "").replace(",", ".")

    # remove sinais
    s = s.replace("+", "").replace("-", "")

    # se não tiver número válido, retorna 0
    if not re.search(r"\d", s):
        return 0.0

    # mantém só números e ponto
    s = re.sub(r"[^0-9.]", "", s)

    if not s:
        return 0.0

    try:
        return float(s)
    except:
        return 0.0


def col_para_indice(col_letra: str) -> int:
    return ord(col_letra.upper()) - ord("A") + 1


def indice_para_col(idx: int) -> str:
    return chr(ord("A") + idx - 1)


# ── 1. Buscar arquivo de entrada ──────────────────────────────

def encontrar_arquivo_entrada() -> Path:
    if not ENTRADAS_DIR.exists():
        raise Exception(f"Pasta não encontrada: {ENTRADAS_DIR}")

    arquivos = sorted([p for p in ENTRADAS_DIR.iterdir() if p.suffix.lower() == ".xlsx"])
    if not arquivos:
        raise Exception("Nenhum arquivo .xlsx encontrado em /entradas")

    return arquivos[0]


# ── 2. Ler extrato XLSX ───────────────────────────────────────

def ler_extrato_xlsx(caminho: Path):
    wb = load_workbook(caminho, data_only=True)
    ws = wb.active

    header_row = None
    col_data = col_hist = col_debito = col_credito = None

    for row in range(1, min(ws.max_row, 30) + 1):
        valores = []
        for col in range(1, min(ws.max_column, 15) + 1):
            val = ws.cell(row, col).value
            valores.append("" if val is None else str(val).strip())

        vals_norm = [norm(v) for v in valores]

        achou_data = None
        achou_hist = None
        achou_debito = None
        achou_credito = None

        for i, v in enumerate(vals_norm):
            if v == "data":
                achou_data = i + 1
            elif "historico" in v:
                achou_hist = i + 1
            elif "debito" in v:
                achou_debito = i + 1
            elif "credito" in v:
                achou_credito = i + 1

        if achou_data and achou_hist and achou_debito:
            header_row = row
            col_data = achou_data
            col_hist = achou_hist
            col_debito = achou_debito
            col_credito = achou_credito
            break

    if not header_row or not col_data or not col_hist or not col_debito:
        raise Exception("Não foi possível localizar cabeçalhos DATA / HISTÓRICO / DÉBITO no extrato.")

    lancamentos = []

    for row in range(header_row + 1, ws.max_row + 1):
        data_val = ws.cell(row, col_data).value
        hist_val = ws.cell(row, col_hist).value
        debito_val = ws.cell(row, col_debito).value

        if data_val is None and hist_val is None and debito_val is None:
            continue

        historico = "" if hist_val is None else str(hist_val).strip()
        if not historico:
            continue

        hist_norm = norm(historico)

        # ignora linhas-resumo / cabeçalhos repetidos
        if hist_norm in ["historico", "conta", "saldo", "credito", "debito"]:
            continue

        if "saldo mes anterior" in hist_norm:
            continue

        debito = limpar_valor_excel(debito_val)
        if debito <= 0:
            continue

        if hasattr(data_val, "strftime"):
            data_str = data_val.strftime("%d/%m/%Y")
            mes = data_val.month
            ano = data_val.year
        else:
            data_str = str(data_val).strip()
            m = re.match(r"(\d{2})/(\d{2})/(\d{4})", data_str)
            if not m:
                continue
            mes = int(m.group(2))
            ano = int(m.group(3))

        lancamentos.append({
            "data": data_str,
            "descricao": historico,
            "valor": abs(debito),
            "mes": mes,
            "ano": ano,
        })

    if not lancamentos:
        raise Exception("Nenhum débito encontrado no arquivo.")

    return lancamentos


# ── 3. Classificação ──────────────────────────────────────────

def classificar_keywords(descricao: str):
    desc = norm(descricao)
    melhor = 999
    score = 0

    for num, info in ROW_MAP.items():
        for kw in info["keywords"]:
            kw_n = norm(kw)
            if kw_n in desc and len(kw_n) > score:
                score = len(kw_n)
                melhor = num

    return melhor, score


def classificar_com_claude(lancamentos):
    print(f"[2/5] Classificando {len(lancamentos)} lançamentos com Claude AI...")
    client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)

    categorias = "\n".join([
        f"Linha {num}: {info['empresa'] + ' — ' if info['empresa'] else ''}{info['desc']} (Keywords: {', '.join(info['keywords'])})"
        for num, info in ROW_MAP.items()
    ])

    lista_formatada = "\n".join([
        f"{i+1}. Data: {l['data']} | Valor: R$ {l['valor']:.2f} | Descrição: '{l['descricao']}'"
        for i, l in enumerate(lancamentos)
    ])

    prompt = f"""Você é o analista financeiro do condomínio Diamond Tower. Classifique cada débito do extrato nas linhas do orçamento.

CATEGORIAS:
{categorias}

Linha 999 = não classificado / dúvida

REGRAS IMPORTANTES:
1. Marco Aurelio = linha 13
2. Subsindico / Pro-labore subsindico = linha 14
3. ONX / portaria / recepção / limpeza terceirizada = linha 4
4. Guarida taxa adm = linha 12
5. Cappelletto Gestão = linha 11
6. FGTS / INSS / ISSQN / SECOVIMED / EFD-REINF = linha 51
7. DMAE = linha 59
8. CEEE = linha 60
9. Seguro / Tokio Marine = linha 61
10. Schindler/Atlas: manutenção mensal = linha 18; peças/conserto = linha 41

LANÇAMENTOS:
{lista_formatada}

Responda SOMENTE em JSON válido:
{{
  "classificacoes": [
    {{"n": 1, "linha": 13, "motivo": "..." }}
  ]
}}
"""

    try:
        resp = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=2500,
            messages=[{"role": "user", "content": prompt}],
        )
        texto = resp.content[0].text.strip()
        texto = re.sub(r"```json\n?|\n?```", "", texto).strip()
        resultado = json.loads(texto)

        classif = {}
        for c in resultado.get("classificacoes", []):
            idx = c["n"] - 1
            if 0 <= idx < len(lancamentos):
                classif[idx] = {"linha": c["linha"], "motivo": c.get("motivo", "")}

        for i, l in enumerate(lancamentos):
            if i not in classif:
                linha, _ = classificar_keywords(l["descricao"])
                classif[i] = {"linha": linha, "motivo": "fallback keywords"}

        return classif

    except Exception as e:
        print(f"   Claude falhou, usando fallback por keywords. Erro: {e}")
        return {
            i: {"linha": classificar_keywords(l["descricao"])[0], "motivo": "fallback geral"}
            for i, l in enumerate(lancamentos)
        }


# ── 4. Atualizar PF.xlsx ──────────────────────────────────────

def adicionar_nota_existente(nota_atual, novas_linhas):
    bloco_novo = "LANÇAMENTOS DA GUARIDA:\n" + "\n".join(novas_linhas)
    if nota_atual and str(nota_atual).strip():
        return str(nota_atual).rstrip() + "\n" + "\n".join(novas_linhas)
    return bloco_novo


def atualizar_pf(lancamentos, classificacoes):
    print("[3/5] Atualizando PF.xlsx...")

    if not PF_PATH.exists():
        raise Exception(f"Arquivo não encontrado: {PF_PATH}")

    wb = load_workbook(PF_PATH)
    anos_planilha = set()

    for l in lancamentos:
        ano_aba = l["ano"] if l["mes"] >= 3 else l["ano"] - 1
        anos_planilha.add(ano_aba)

    if len(anos_planilha) != 1:
        raise Exception("O arquivo de entrada possui mais de um período-base de PF.")
    ano_base = list(anos_planilha)[0]

    aba_nome = f"PF {ano_base}"
    if aba_nome not in wb.sheetnames:
        raise Exception(f"Aba não encontrada no PF.xlsx: {aba_nome}")

    ws = wb[aba_nome]

    por_celula = {}
    nao_classificados = []

    for i, lanc in enumerate(lancamentos):
        linha = classificacoes.get(i, {}).get("linha", 999)

        if 5 <= linha <= 10:
            linha = 4

        if linha == 999 or linha not in ROW_MAP:
            nao_classificados.append(lanc)
            continue

        col_letra = MONTH_TO_COL.get((lanc["mes"], lanc["ano"]))
        if not col_letra:
            nao_classificados.append(lanc)
            continue

        celula = f"{col_letra}{linha}"

        if celula not in por_celula:
            por_celula[celula] = {
                "total": 0.0,
                "itens": [],
                "desc": ROW_MAP[linha]["desc"],
            }

        por_celula[celula]["total"] += lanc["valor"]
        por_celula[celula]["itens"].append(
            f"{lanc['data']} - {lanc['descricao']} | R$ {formatar_brl(lanc['valor'])}"
        )

    resumo = []

    for cel_ref, dados in por_celula.items():
        cell = ws[cel_ref]

        valor_atual = cell.value
        if valor_atual in (None, ""):
            valor_atual_num = 0.0
        elif isinstance(valor_atual, (int, float)):
            valor_atual_num = float(valor_atual)
        else:
            try:
                valor_atual_num = limpar_valor_excel(valor_atual)
            except Exception:
                valor_atual_num = 0.0

        novo_valor = round(valor_atual_num + dados["total"], 2)
        cell.value = novo_valor
        cell.number_format = '#,##0.00'

        nota_atual = cell.comment.text if cell.comment else ""
        nota_final = adicionar_nota_existente(nota_atual, dados["itens"])

        from openpyxl.comments import Comment
        cell.comment = Comment(nota_final, "Automação")

        resumo.append({
            "celula": cel_ref,
            "desc": dados["desc"],
            "valor": novo_valor,
            "itens": dados["itens"],
        })

    wb.save(PF_PATH)

    return resumo, nao_classificados


# ── 5. Mover para processados ─────────────────────────────────

def mover_para_processados(arquivo: Path):
    PROCESSADOS_DIR.mkdir(parents=True, exist_ok=True)
    destino = PROCESSADOS_DIR / arquivo.name

    if destino.exists():
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        destino = PROCESSADOS_DIR / f"{arquivo.stem}_{timestamp}{arquivo.suffix}"

    shutil.move(str(arquivo), str(destino))
    return destino


# ── MAIN ──────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print(" DIAMOND TOWER — Automação Orçamentária via XLSX")
    print(f" {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("=" * 60)

    arquivo = encontrar_arquivo_entrada()
    print(f"[1/5] Arquivo encontrado: {arquivo.name}")

    lancamentos = ler_extrato_xlsx(arquivo)
    print(f"   Débitos encontrados: {len(lancamentos)}")

    meses = sorted({(l['mes'], l['ano']) for l in lancamentos})
    print(f"   Competências no arquivo: {meses}")

    classificacoes = classificar_com_claude(lancamentos)

    resumo, nao_classificados = atualizar_pf(lancamentos, classificacoes)

    destino = mover_para_processados(arquivo)
    print(f"[4/5] Arquivo movido para: {destino}")

    print("[5/5] Concluído.\n")

    for r in resumo:
        print(f"{r['celula']:>6} | {r['desc'][:45]:45s} | R$ {formatar_brl(r['valor'])}")

    if nao_classificados:
        print(f"\n⚠ Não classificados: {len(nao_classificados)}")
        for l in nao_classificados[:20]:
            print(f" - {l['data']} | {l['descricao'][:70]} | R$ {formatar_brl(l['valor'])}")


if __name__ == "__main__":
    main()
